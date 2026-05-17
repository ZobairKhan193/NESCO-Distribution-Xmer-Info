"""
Dump the structure of every source Excel file in the data workspace so we can
design converters from there. Prints sheet names, dimensions, first 3 header
rows, and a few sample rows per sheet.

Run from repo root (or anywhere — paths are absolute to the workspace).
"""
import io
import sys
from pathlib import Path

# Force UTF-8 stdout so cell text with Bangla/curly quotes does not crash on
# Windows code page 1252.
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]   # f:/NESCO Web Design

FILES = [
    ROOT / "33-11 kV SS Info" / "All 33-11 kV Substation Info Master File.xlsx",
    ROOT / "33-11 kV SS Info" / "33-11 kV Substation Information Template.xlsx",
    ROOT / "Distribution Transformer" / "(With Dashboard) Summary_from_all_the_SDD-ESU.xlsx",
    ROOT / "NIDMP" / "ADB_AIS_Substation_Upgradation_Summary (Claud).xlsx",
    ROOT / "NIDMP" / "Grid_Bay_Breakers.xlsx",
    ROOT / "PDSSP" / "NDB SS_Upgradation_Substation_Summary.xlsx",
    ROOT / "PDSSP" / "Line Requirement for All SDDS (Updated).xlsx",
    ROOT / "Switching Substations Info" / "Grid Substation Info.xlsx",
    ROOT / "Switching Substations Info" / "Grid or Switching SS List (with Google Map Location).xlsx",
    ROOT / "Switching Substations Info" / "switching substation info.xlsx",
    ROOT / "Store" / "NESCO_Substation_Equipment_List.xlsx",
    ROOT / "Store" / "NESCO_Line_Equipment_List.xlsx",
    ROOT / "ZRS" / "ZRS_Transformer_Consolidated_2025.xlsx",
    ROOT / "Technical_Highlights foir Homepage.xlsx",
]


def short(v, n=30):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").strip()
    return s if len(s) <= n else s[: n - 1] + "…"


def dump(path: Path):
    if not path.exists():
        print(f"\n!!! MISSING: {path}")
        return
    print(f"\n=== {path.relative_to(ROOT)} ({path.stat().st_size:,} bytes) ===")
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        print(f"  (cannot open: {e})")
        return
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  --- sheet [{sn}]  {ws.max_row} rows × {ws.max_column} cols ---")
        # first 5 rows
        for r in range(1, min(ws.max_row, 6) + 1):
            cells = [short(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 12) + 1)]
            print("  R%2d | %s" % (r, " | ".join(cells)))
        if ws.max_row > 6:
            print(f"  ... ({ws.max_row - 6} more rows)")


def main():
    for f in FILES:
        dump(f)


if __name__ == "__main__":
    main()
